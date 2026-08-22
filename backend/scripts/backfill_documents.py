# backend/scripts/backfill_documents.py
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pathlib import Path
from app.supabase_client import supabase
import fitz  # PyMuPDF
import docx
import openpyxl
from PIL import Image
import pytesseract

BASE_DIR = Path(__file__).resolve().parents[1]
UPLOAD_DIR = BASE_DIR / "uploads"

def extract_metadata(file_path: Path, ext: str):
    """Extract text and metadata from document"""
    chunks = []
    extracted_text = ""
    
    try:
        if ext == ".pdf":
            doc = fitz.open(str(file_path))
            for i, page in enumerate(doc):
                text = page.get_text()
                extracted_text += text
                chunks.append({
                    "id": f"backfill#p{i+1}s1",
                    "page": i + 1,
                    "section": f"Page {i+1}",
                    "text": text.strip()
                })
            doc.close()
            
        elif ext == ".docx":
            doc = docx.Document(str(file_path))
            text = "\n".join([para.text for para in doc.paragraphs])
            extracted_text = text
            chunks.append({"id": "backfill#p1s1", "page": 1, "section": "Document Body", "text": text.strip()})
            
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(str(file_path), read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_text = []
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    sheet_text.append(row_text)
                text = "\n".join(sheet_text)
                extracted_text += f"--- Sheet: {sheet} ---\n{text}\n"
                chunks.append({"id": f"backfill_sheet_{sheet}", "page": 1, "section": f"Sheet: {sheet}", "text": text.strip()})
            wb.close()
            
        elif ext in [".jpg", ".jpeg", ".png"]:
            try:
                text = pytesseract.image_to_string(Image.open(str(file_path)))
                extracted_text = text
                chunks.append({"id": "backfill#p1s1", "page": 1, "section": "OCR Extracted", "text": text.strip()})
            except:
                chunks.append({"id": "backfill#p1s1", "page": 1, "section": "Image", "text": "[Image file]"})
        else:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
                extracted_text = text
                chunks.append({"id": "backfill#p1s1", "page": 1, "section": "Content", "text": text.strip()})
                
        return chunks, extracted_text
        
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return [], ""

def backfill_documents():
    """Process all documents with NULL metadata"""
    
    # Get all documents that need processing
    result = supabase.table("documents").select("*").execute()
    documents = result.data
    
    print(f"Found {len(documents)} documents to process")
    
    for doc in documents:
        doc_id = doc["id"]
        file_name = doc["file_name"]
        status = doc.get("status", "uploaded")
        
        # Skip if already indexed
        if status == "Indexed" and doc.get("chunks"):
            print(f"✓ Skipping {file_name} (already indexed)")
            continue
        
        ext = Path(file_name).suffix.lower()
        save_path = UPLOAD_DIR / f"{doc_id}{ext}"
        
        if not save_path.exists():
            print(f"✗ File not found: {save_path}")
            continue
        
        print(f"Processing: {file_name}")
        
        # Extract metadata
        chunks, extracted_text = extract_metadata(save_path, ext)
        
        # Update document
        update_data = {
            "status": "Indexed",
            "chunks": chunks,
            "confidence": 0.94 if chunks else 0.0,
            "tags": ["backfilled", "auto-indexed", ext.replace(".", "")],
            "format": ext.replace(".", "").upper() if ext else None,
        }
        
        # Try to infer title from filename
        if not doc.get("title"):
            update_data["title"] = Path(file_name).stem
        
        try:
            supabase.table("documents").update(update_data).eq("id", doc_id).execute()
            print(f"✓ Indexed {file_name} ({len(chunks)} chunks)")
        except Exception as e:
            print(f"✗ Failed to update {file_name}: {e}")

if __name__ == "__main__":
    backfill_documents()